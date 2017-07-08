import openpyxl
doc = openpyxl.load_workbook('nomina_mayo_2017/MUNICIPIO DE GUADALAJARA_5_2017.xlsx')

hoja = doc.get_sheet_by_name('2017_5')

#print hoja.max_row
#rango hoja['T2:T32019']
#print hoja['T'].value
#columna = hoja['T']

wb = openpyxl.load_workbook("nomina_mayo_2017/nomina_mayo_52.xlsx")
ws1 = wb.get_sheet_by_name("nomina_listado_uuid ")
columnaB = ws1['I8:I9']
b = 2
mylist = []
for a in xrange(32019):
   mylist.append( hoja['T'+ str(b)].value )
   b = b + 1        

y = 8
rango = len(columnaB)
for x in xrange(rango): 
   ws1.cell(row = y, column = 18, value = str(ws1['I'+ str(y)].value  in  mylist))
   y = y + 1
  # if x == 60 : 
    # break

wb.save("nomina_mayo_52.xlsx")
